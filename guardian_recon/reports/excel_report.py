"""
GuardianRecon — Excel Report Generator
==========================================
يصدّر نتيجة التسوية إلى ملف Excel احترافي:
  - ورقة الملخص (Standard bank rec format)
  - ورقة العناصر المطابقة
  - ورقة العناصر التي تحتاج قيد تسوية
  - ورقة العناصر التي تحتاج تحقيق (مع تلوين حسب التقادم)
"""

from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..engine.models import ReconCategory, ItemStatus

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
BOLD = Font(bold=True)

STATUS_COLORS = {
    ItemStatus.CURRENT: "C6EFCE",    # أخضر
    ItemStatus.AGING: "FFEB9C",      # أصفر
    ItemStatus.OVERDUE: "FFC7CE",    # وردي/أحمر فاتح
    ItemStatus.STALE: "FF7C7C",      # أحمر
}

THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def _style_header_row(ws, row_idx: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autofit(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def build_summary_sheet(wb: Workbook, summary: dict):
    ws = wb.active
    ws.title = "الملخص"
    ws.sheet_view.rightToLeft = True

    ws["B2"] = "تقرير التسوية البنكية — GuardianRecon"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"كما في تاريخ: {summary['as_of']}"

    rows = [
        ("رصيد كشف الحساب البنكي", summary["bank_balance"]),
        ("رصيد دفتر الأستاذ (GL)", summary["gl_balance"]),
        ("الفرق الخام", summary["raw_difference"]),
        ("", ""),
        ("إجمالي العناصر", summary["total_items"]),
        ("عناصر مطابقة تماماً", summary["matched_count"]),
        ("فروقات توقيت (Timing)", summary["timing_diff_count"]),
        ("تحتاج قيد تسوية (Adjustment)", summary["adjustment_count"]),
        ("تحتاج تحقيق (Investigation)", summary["investigation_count"]),
        ("إجمالي العناصر غير المسوّاة", summary["outstanding_total"]),
        ("عناصر تحتاج تصعيد (Escalation)", len(summary["escalations"])),
    ]

    r = 5
    for label, value in rows:
        ws.cell(row=r, column=2, value=label).font = BOLD
        ws.cell(row=r, column=3, value=value)
        r += 1

    _autofit(ws, {"A": 3, "B": 35, "C": 20})


def build_items_sheet(wb: Workbook, title: str, items: list):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True

    headers = ["التاريخ", "المصدر", "الوصف", "المبلغ", "الفرق",
               "الحالة", "التقادم (يوم)", "نسبة الثقة", "ملاحظة"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for item in items:
        source = "بنك" if item.bank_txn and not item.gl_txn else \
                 "GL" if item.gl_txn and not item.bank_txn else "بنك+GL"
        ref_txn = item.bank_txn or item.gl_txn
        desc = getattr(ref_txn, "description_en", None) or getattr(ref_txn, "description", "") if ref_txn else ""
        amount = ref_txn.amount if ref_txn else 0
        txn_date = ref_txn.txn_date if ref_txn else ""

        row = [
            txn_date, source, desc, amount, item.difference,
            item.status.value, item.age_days, item.match_confidence, item.note,
        ]
        ws.append(row)

        status_color = STATUS_COLORS.get(item.status)
        if status_color and item.category != ReconCategory.MATCHED:
            fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    _autofit(ws, {"A": 12, "B": 10, "C": 45, "D": 14, "E": 12,
                   "F": 14, "G": 12, "H": 12, "I": 45})


def export_reconciliation_report(summary: dict, output_path: str):
    wb = Workbook()
    build_summary_sheet(wb, summary)

    build_items_sheet(wb, "مطابقة تماماً",
                       summary["items_by_category"][ReconCategory.MATCHED.value])
    build_items_sheet(wb, "فروقات توقيت",
                       summary["items_by_category"][ReconCategory.TIMING_DIFFERENCE.value])
    build_items_sheet(wb, "تحتاج قيد تسوية",
                       summary["items_by_category"][ReconCategory.ADJUSTMENT_REQUIRED.value])
    build_items_sheet(wb, "تحتاج تحقيق",
                       summary["items_by_category"][ReconCategory.REQUIRES_INVESTIGATION.value])

    if summary["escalations"]:
        build_items_sheet(wb, "تصعيد ⚠️", summary["escalations"])

    wb.save(output_path)
    return output_path
